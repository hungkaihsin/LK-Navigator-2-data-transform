import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
import matplotlib.pyplot as plt 
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter import colorchooser as cc
from tkinter.messagebox import showinfo

# Main root
root = tk.Tk()
root.title("Visualisation")
root.resizable(False, False)
root.geometry("400x400")
root.iconbitmap("logo.ico")


# Label
lb_choose_file = ttk.Label(text = "Choose file:")
lb_choose_file.place(x= 0, y = 3)
lb_measuredata = ttk.Label(text = "Measure data:")
lb_measuredata.place(x = 0, y = 37)
lb_measuredata_remind = ttk.Label(text = "(depend on amount data)")
lb_measuredata_remind.place(x=200, y = 37 )
lb_time = ttk.Label(text="Measure time:")
lb_time.place(x = 0, y = 77)
lb_time_remind = ttk.Label(text = "(s)")
lb_time_remind.place(x = 200, y= 77)
lb_color_choosen = tk.Label(height=1, width=2, bg= "green")
lb_color_choosen.place(x=195, y= 111)




# Entry
en_choose_file = ttk.Entry(width= 25)
en_choose_file.place(x = 85, y = 0)
en_measuredate = ttk.Entry(width= 8)
en_measuredate.place(x = 100, y = 35)
en_time = ttk.Entry(width= 8)
en_time.place(x= 100, y = 75)


# Function

# Color chooser
def colorchoose():
    color = cc.askcolor()
    color = color[1]
    print(str(color))
    lb_color_choosen.config(bg= color)

def load_file():
    filetypes = (
        ('csv file', '*.csv'),('All files', '*.*')
    )
    if en_choose_file.get() is None:
        filepath = fd.askopenfilename(
        title = 'Open a file',
        initialdir = '/',
        filetypes = filetypes)
    else:
        filepath = fd.askopenfilename(
        title = 'Open a file',
        initialdir = '/',
        filetypes = filetypes)
        en_choose_file.delete(0,'end')
        en_choose_file.insert(0, filepath)

def transit():
    if len(en_choose_file.get()) == 0:
        showinfo(title = 'Wrong', message = 'Please choose file')
    elif len(en_measuredate.get()) == 0:
        showinfo(title = 'Wrong', message = 'Please type data')
    elif len(en_time.get()) == 0:
        showinfo(title = 'Wrong', message = 'Please type time')
    else:
        file_route = en_choose_file.get()
        measure_data = en_measuredate.get()
        measure_data = int(measure_data)
        measure_time = en_time.get()
        measure_time = int(measure_time)
        transition = float(measure_time / measure_data)

        load_csv_file = pd.read_csv(file_route, usecols=[2])
        header_list = ["Distance"]
        load_csv_file.to_excel("123456.xlsx", header= header_list, index= False)


        wb_open = openpyxl.load_workbook("123456.xlsx")
        sheet = wb_open.worksheets[0]
        for transit in range(1, measure_data, 1):
            count = transit
            outcome=  count * transition
            sheet.cell((transit + 1),2).value = outcome

        
        sheet.cell(1,2).value = "Time"
        sheet.cell(1,2).font = Font(bold = True)
        sheet["B1"].alignment = Alignment(horizontal= "center")
        wb_open.save("123456.xlsx")
        showinfo(
        title = "Transform",
        message = "Done!",
    )


        read_excel = pd.read_excel("123456.xlsx")
        x_axis = read_excel["Time"]
        y_axis = read_excel["Distance"]
        
        plt.plot(x_axis, y_axis, color = "#000000")
        plt.xlabel("Time(s)", fontsize = "12")
        plt.ylabel("Distance(mm)", fontsize = "12")
        plt.title("Actuator", fontsize = "15")
        plt.ylim(-7,2)
        plt.show()









# Button
btn_loadfile = ttk.Button(text = '...', command= load_file, width = 4)
btn_loadfile.place(x = 330, y = 0, width= 50)
btn_transit = ttk.Button(text = "Transit", command= transit)
btn_transit.place(x = 150, y = 170)
btn_choose_color = ttk.Button(text="Choose plot's line color", command= colorchoose)
btn_choose_color.place(x=0,y= 109)




root.mainloop()



